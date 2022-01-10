# 엑셀 읽기
from openpyxl import load_workbook

wb = load_workbook("./sample2.xlsx")

ws1 = wb.active

# cell 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws1.cell(x, y).value, end=" ")
#     print()

# max_row / max_column : 시트가 가지고 있는 행, 열 갯수 가져오기
for x in range(1, ws1.max_row + 1):
    for y in range(1, ws1.max_column + 1):
        print(ws1.cell(x, y).value, end=" ")
    print()

wb.close()

# 셀
from openpyxl import Workbook
import random

wb = Workbook()

# 기본시트 활성화
ws1 = wb.active
# 시트명 변경
ws1.title = "시트1"

# A1 셀에 1이라는 값을 입력
ws1["A1"] = 1
ws1["A2"] = 2
ws1["A3"] = 3

ws1["B1"] = 1
ws1["B2"] = 2
ws1["B3"] = 3

# 셀에 접근해서 값 가져오기
print(ws1["A1"])  # <Cell '시트1'.A1> => A1 셀 정보 출력
print(ws1["A2"].value)  # 2 => A2 셀 값 출력

# 셀을 접근하는 두번째 방법
# row = 1, 2, 3..., column = 1(A), 2(B), 3(C)...
print(ws1.cell(1, 1).value)  # ws["A1"].value
print(ws1.cell(1, 2).value)  # ws["B1"].value

c = ws1.cell(row=1, column=3, value=10)
print(c.value)

# for 반복문을 통해서 랜덤으로 숫자 채우기
for x in range(1, 11):
    for y in range(1, 11):
        ws1.cell(row=x, column=y, value=random.randint(0, 100))

wb.save("sample2.xlsx")
wb.close()

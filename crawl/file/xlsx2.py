# 시트
from openpyxl import Workbook

wb = Workbook()

# 새로운 시트 생성하기
ws1 = wb.create_sheet()
# 생성된 시트 이름 변경
ws1.title = "Mysheet"
# 시트 탭 색상 변경
ws1.sheet_properties.tabColor = "ff66ff"

# 두 번째 시트 생성(시트명, 위치)
ws2 = wb.create_sheet("급여명세", 2)

# 시트 가져오기
new_ws = wb["Mysheet"]

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

# 현재 생성된 모든 시트 이름 출력
print(wb.sheetnames)

# 저장
wb.save("sample1.xlsx")

wb.close()

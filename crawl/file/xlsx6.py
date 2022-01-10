from openpyxl import load_workbook

wb = load_workbook("range.xlsx")

ws1 = wb.active

# 영어 점수가 80점 이상인 학생의 점수
for row in ws1.iter_rows(min_row=2):  # 제목 행 제외
    if int(row[1].value) >= 80:
        print(row[0].value, " 번 학생의 영어 점수는 ", row[1].value, " 입니다.")

# 영어 과목명 변경 => 컴퓨터
for row in ws1.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value == "컴퓨터"

wb.save("range_modified.xlsx")

wb.close()

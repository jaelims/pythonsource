# 셀에 데이터 삽입
from openpyxl import workbook
from random import randint, randrange

from openpyxl.workbook.workbook import Workbook

wb = Workbook()

ws1 = wb.active

# 1줄씩 데이터 넣기
ws1.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws1.append([i, randint(0, 100), randint(0, 100)])

# 특정과목만 가져오기 - 영어
# col_B = ws1["B"]
# # print(col_B)
# for cell in col_B:
#     print(cell.value)

# # 영어, 수학 두 개 가져오기
# col_range = ws1["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value, end=" ")
#     print()

# 1번 row 가져오기
# row_title = ws1[1]
# for cell in row_title:
#     print(cell.value)

# 2 ~ 5 번 ROW 가져오기
# row_range = ws1[2:5]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# 3행부터 끝까지 가져오기
# row_range = ws1[3 : ws1.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# iter_rows() : 전체 rows를 가져오는 함수
for row in ws1.iter_rows():
    print(row[2].value)

# wb.save("range.xlsx")

wb.close()
